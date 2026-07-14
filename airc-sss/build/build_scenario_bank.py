# -*- coding: utf-8 -*-
"""
Builds Maricopa_Scenario_Bank.xlsx — the data backbone for the scaled
Domain 5 Student Support & Success usability study.

Model: every test is ONE persona + ONE campus + ONE task. The persona gets a
problem, never a destination. Base tasks multiply by persona and by the 10
colleges to reach the ~500-scenario / ~5,000-task-run goal.

No PII. Personas are research instruments mapped to test accounts. Curly
quotes only, no em dashes.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette (matches sss.css) ----
PLUM="7A5080"; PLUM_PALE="EDE4F2"; SAGE="6B8F6E"; SAGE_PALE="E4EDE5"
GOLD="B8956A"; GOLD_PALE="FDF6EE"; ROSE="C4929E"; ROSE_PALE="F5E8EB"
INK="3A2E3F"; STONE="E2D8CE"; PAPER="FAF7F2"

def fill(hex): return PatternFill("solid", fgColor=hex)
def F(sz=10,b=False,color=INK,it=False): return Font(name="Calibri",size=sz,bold=b,color=color,italic=it)
thin=Side(style="thin",color=STONE)
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")
TOP=Alignment(vertical="top")
CTR=Alignment(horizontal="center",vertical="top")

# =====================================================================
# 1. THE TEN COLLEGES
# =====================================================================
COLLEGES=[
 ("CGCC","Chandler-Gilbert Community College"),
 ("EMCC","Estrella Mountain Community College"),
 ("GateWay","GateWay Community College"),
 ("GCC","Glendale Community College"),
 ("MCC","Mesa Community College"),
 ("PVCC","Paradise Valley Community College"),
 ("Phoenix","Phoenix College"),
 ("Rio","Rio Salado College"),
 ("SCC","Scottsdale Community College"),
 ("SMCC","South Mountain Community College"),
]

# =====================================================================
# 2. PERSONA LIBRARY  (extends the original 3: Marisol, Darnell, Alex)
#    fields: id, name, age, background, barriers, goal, tech, language, campuses
# =====================================================================
PERSONAS=[
 # --- the three original, carried forward unchanged ---
 ("P01","Marisol Reyes",19,"First-gen, pre-Nursing/CTE, works part-time at a grocery store, lives with family.",
  "Spanish at home, reads English well but second-guesses forms; Pell-dependent; phone-first, no laptop; never been advised.",
  "Get fully enrolled and find money for the first semester without missing a deadline.","Phone-first, low confidence with portals","Spanish","GCC; swirls to EMCC"),
 ("P02","Darnell Carter",34,"Veteran, Business AA bound for ASU transfer, full-time job, two kids.",
  "Impatient with slow processes; self-advises; needs VA benefits to certify; limited daytime hours; never sees an advisor.",
  "Use his GI Bill benefits and transfer cleanly to ASU without losing credits.","High; expects fast self-service","English","SCC; swirls to Rio"),
 ("P03","Alex Nguyen",18,"Exploring/undecided, leaning digital media, just out of high school, lives at home.",
  "ADHD and anxiety; needs DRS accommodations; institutionally naive; wants to belong; abandons tasks when overwhelmed.",
  "Pick a direction, set up accommodations, and feel like they belong before week one.","Digitally fluent, institutionally naive","English","GCC"),
 # --- new personas ---
 ("P04","Ana Lucia Torres",27,"ESL adult learner, came through Adult Education / GED, wants an allied-health certificate.",
  "Spanish primary; limited English on dense pages; works nights cleaning offices; no childcare; afraid of owing money.",
  "Move from GED into a short certificate and understand what it costs before she commits.","Low; uses phone and translate","Spanish","Phoenix; Rio for one online class"),
 ("P05","Hassan Abdi",22,"Refugee background, English is a third language, first in family at U.S. college.",
  "Arabic and Somali at home; unfamiliar with U.S. registration; relies on a relative to interpret forms; phone data limited.",
  "Apply, get an MEID, and register for ESL and gen-ed without coming to campus.","Low to moderate; phone-first","Somali","EMCC"),
 ("P06","Robert Kessler",41,"Returning adult learner, laid off from a trade, retraining for IT.",
  "Hasn't been in school in 20 years; no recent transcripts on hand; unsure if old credits count; evening-only availability.",
  "Find out what transfers in and build a part-time evening schedule he can keep.","Moderate; desktop at the library","English","MCC; swirls to SCC"),
 ("P07","Jasmine Williams",24,"Single working parent, business major, food insecure, drives 35 minutes to campus.",
  "Skips meals; doesn't know basic-needs help exists; juggles a toddler; no margin for a wasted trip; misses email.",
  "Find food and emergency help and keep her schedule without dropping below full-time aid.","Moderate; phone and text","English","CGCC"),
 ("P08","Kevin Nguyen",17,"Dual-enrollment high schooler taking college classes at his high school.",
  "Minor, parent involved; juggles high-school and college calendars; unclear which login is which; first college experience.",
  "Register for a dual-enrollment class and find his college email and Canvas.","High; school-issued device","English","GateWay; ACE program"),
 ("P09","Linda Yazzie",38,"American Indian student, commutes from a tribal community, accounting major.",
  "Long drive, spotty rural broadband; needs American Indian Program support; works seasonally; cautious about asking for help.",
  "Connect with the American Indian Program and get tutoring she can use from home.","Low bandwidth; phone-first","English","SCC"),
 ("P10","Tyler Brooks",20,"Autistic student, computer science, lives at home, very capable academically.",
  "Needs clear literal instructions; sensory overload on busy pages; phone calls are hard; needs DRS; rigid routines.",
  "Set up accommodations and register entirely online without phone calls or a campus visit.","High; prefers written channels","English","PVCC"),
 ("P11","Destiny Carter",19,"Foster-care alumna, no family safety net, undecided, Pell-dependent.",
  "Housing-insecure; needs emergency aid fast; no one to ask; distrusts institutions; can lose contact during moves.",
  "Find emergency housing and money help and stay enrolled through a crisis.","Moderate; phone-first","English","SMCC"),
 ("P12","Mei Chen",30,"International student on an F-1 visa, accounting, away from family.",
  "Visa rules limit work hours; unfamiliar with U.S. aid (ineligible for federal aid); needs international-student office.",
  "Keep her visa status valid while registering and paying without federal aid.","High; laptop","Mandarin","SCC"),
 ("P13","Marcus Johnson",26,"Justice-involved (reentry), warehouse job, wants a logistics certificate.",
  "Background gaps in records; unsure if past affects admission; limited time; needs clear nonjudgmental info.",
  "Apply without fear of being turned away and find a short, job-aligned certificate.","Moderate; phone","English","SMCC"),
 ("P14","Priya Patel",22,"Transfer-bound honors student, biology, high achiever, anxious about credits.",
  "Perfectionist; terrified of taking a class that won't transfer; juggles a campus job; over-researches everything.",
  "Build a transfer plan to a university with zero wasted credits and join Honors.","High; laptop","English","CGCC; swirls to MCC"),
 ("P15","Frank DiMarco",58,"Career-changer near retirement, wants a real-estate or coding certificate.",
  "Low digital confidence; intimidated by portals; needs a human early; can't find the help-desk number.",
  "Get help from a real person to enroll in one certificate without getting lost online.","Low; needs phone support","English","PVCC"),
 ("P16","Sofia Ramirez",18,"DACA / Dreamer, first-gen, hospitality major, unsure of her aid options.",
  "Ineligible for federal aid, eligible for some state/private; afraid to disclose status; needs accurate, safe info.",
  "Find scholarships she actually qualifies for without risking her status.","Moderate; phone-first","Spanish","Phoenix"),
 ("P17","DeShawn Allen",21,"Student-athlete, kinesiology, heavy practice schedule, eligibility-driven.",
  "Must stay full-time for eligibility; tight time; needs to know drop/withdraw impact before he touches a class.",
  "Drop one class without losing eligibility or aid and replace it the same week.","Moderate; phone","English","GCC"),
 ("P18","Grace Okafor",45,"Working RN returning for a bachelor-completion path, swirls across campuses.",
  "Time-starved; takes one needed class at whichever campus offers it; consortium aid confusion; residency worry.",
  "Take one required class at a second campus and have aid and credit follow her.","High; laptop","English","Rio + GCC (swirl)"),
 ("P19","Caleb Smith",19,"Rural student, very low bandwidth, agriculture/business, drives in weekly.",
  "Slow internet at home; can't stream video orientations; needs lightweight, low-data paths; phone data caps.",
  "Complete onboarding and orientation on a slow connection without a campus trip.","Low bandwidth; phone-first","English","EMCC"),
 ("P20","Brittany Hale",23,"First-gen ADHD student, marketing, easily lost in multi-step processes.",
  "Loses the thread on long forms; forgets deadlines; needs reminders; abandons when a step is unclear.",
  "Finish a multi-step financial-aid verification before the deadline without giving up.","Moderate; phone and laptop","English","MCC"),
 ("P21","Omar Haddad",28,"Veteran with a service-connected disability, engineering, needs DRS and VA.",
  "Coordinates DRS accommodations AND VA certification at once; PTSD, avoids crowded offices; needs written steps.",
  "Set up both accommodations and VA benefits through online and written channels.","Moderate; prefers written","Arabic","SCC"),
 ("P22","Nicole Tran",31,"Part-time evening student, full-time medical-biller job, billing/coding.",
  "Only free after 6 PM; offices closed; relies on websites and chat; can't take calls during work.",
  "Do everything after hours: register, pay, and get help when offices are closed.","High; laptop after work","Vietnamese","GateWay"),
 ("P23","Jordan Lee",20,"LGBTQ+ student seeking an affirming campus and counseling support.",
  "Wants to know counseling and community are safe and confidential before reaching out; cautious; reads carefully.",
  "Find affirming counseling and community and confirm it is confidential.","High; phone and laptop","English","PVCC"),
 ("P24","Aaliyah Robinson",19,"Pell-dependent first-gen, undecided, watching every dollar.",
  "Refund timing decides whether she can buy books; doesn't know disbursement dates; can't float costs.",
  "Understand exactly when aid disburses and when her refund lands before classes start.","Moderate; phone-first","English","SMCC"),
 ("P25","Gabriel Mendez",33,"Spanish-dominant parent enrolling after years of work, basic-needs eligible.",
  "Reads forms slowly in English; childcare gap; food insecure; embarrassed to ask; trusts in-person more than web.",
  "Enroll, find childcare and food help, and pay in a language he is comfortable in.","Low; phone with help","Spanish","Phoenix"),
 ("P26","Hannah Berg",18,"Recent grad of a tiny rural high school, undecided, never used an LMS.",
  "Doesn't know what an LMS is; expects email like high school; unsure where classes 'are' online.",
  "Figure out that Canvas is where her classes live and log in before week one.","Moderate; phone","English","CGCC"),
 ("P27","Victor Nguyen",24,"Swirling consortium student taking one online class at a second college.",
  "Already enrolled at home college; needs one class elsewhere; confused by two logins, two aid offices.",
  "Take one consortium class at a second college and have credit return home.","High; laptop","English","Rio + SCC (swirl)"),
 ("P28","Maria Gonzalez",36,"Returning learner, food-insecure single parent, ESL, healthcare aspirations.",
  "Spanish primary; two jobs; no childcare; food pantry need; deeply deadline-anxious; phone-only.",
  "Enroll in a healthcare certificate and line up food and childcare on a phone, in Spanish.","Low; phone-only","Spanish","GateWay"),
 ("P29","Ethan Park",17,"Dual-enrollment high-schooler exploring engineering, ambitious.",
  "Minor; doesn't know college transcripts differ from high-school; unsure how credits carry to a university later.",
  "Take dual-enrollment engineering and understand how the credit follows him.","High; school device","Korean","CGCC; ACE"),
 ("P30","Rosa Jimenez",40,"Adult-ed to college bridge, very low English literacy, cleaning-industry worker.",
  "Minimal English reading; relies on visuals and translate; intimidated by any text-heavy page; cash economy.",
  "Move from Adult Ed into a college certificate with help she can understand.","Very low; needs in-person","Spanish","SMCC"),
 ("P31","Tyrone Davis",29,"Working parent, evening student, needs reliable campus tech and printing.",
  "Relies on campus labs and printing; hits login/DUO walls; wi-fi drops; can't print a needed form.",
  "Print and submit a required form using campus tech without an IT runaround.","Moderate; campus labs","English","MCC"),
 ("P32","Leila Hassan",19,"First-gen Muslim student, business, commutes, needs quiet prayer space and DRS.",
  "Needs to know facilities and accommodations exist; modest about asking; juggles family expectations.",
  "Confirm campus accommodations and quiet space exist before committing.","Moderate; phone","Arabic","GCC"),
 ("P33","Cody Martin",22,"Rural veteran, low bandwidth, agriculture, GI Bill, drives 50 minutes.",
  "Low data; VA certification confusion; can't stream; wants to minimize trips; impatient with portals.",
  "Certify VA benefits and register on a slow connection with minimal campus trips.","Low bandwidth; phone","English","EMCC"),
 ("P34","Samantha Wright",26,"Graduating student offboarding her account, needs to save her work.",
  "About to lose her college email, Drive, and coursework; doesn't know the account closes; needs transcripts too.",
  "Graduate, save her email/Drive/coursework, and order transcripts before the account closes.","High; laptop","English","SCC"),
 ("P35","Andre Thompson",20,"First-gen, undecided, never assigned an advisor, drifting.",
  "No advisor shown in his SIS; doesn't know he can get one; picks classes blindly; at persistence risk.",
  "Find or get assigned an advisor and book a real appointment.","Moderate; phone","English","SMCC"),
 ("P36","Yuki Tanaka",23,"International transfer student, computer science, F-1, prior credits abroad.",
  "Foreign transcript evaluation; visa + registration at once; unsure which credits count; time-zone support gaps.",
  "Get foreign credits evaluated and register while keeping F-1 status.","High; laptop","Japanese","CGCC"),
]

# =====================================================================
# 3. TASK TAXONOMY  — base tasks across the whole student journey
#    fields: code, stage, task (the felt need, problem-not-destination),
#            suspected barrier, needs_salesforce (Y/N)
# =====================================================================
TASKS=[
 # Stage 1 — Get in the door
 ("T01","Get in the door","Apply to the college as a brand-new student.","Multiple entry points, unclear which form, account vs application confusion.","N"),
 ("T02","Get in the door","Get your MEID and first login without coming to campus.","MEID lookup buried, DUO setup hurdle, no obvious help-desk path.","N"),
 ("T03","Get in the door","Find out your residency classification and what it means for tuition.","Residency rules dense, classification status hard to find or contest.","N"),
 ("T04","Get in the door","Complete new-student orientation / FYE.","Orientation hidden or video-only, no low-bandwidth path.","N"),
 ("T05","Get in the door","Register as a dual-enrollment / ACE high-school student.","Which login is which, parent step unclear, separate dual-enroll path.","N"),
 ("T06","Get in the door","Enter through Adult Ed / GED / ESL and find the next step into credit classes.","Adult-Ed-to-college bridge invisible, no Spanish path.","N"),
 ("T07","Get in the door","Get a student ID card.","Where and how to get an ID is unstated, in-person assumed.","N"),
 # Stage 2 — Pay for it
 ("T08","Pay for it","Apply for financial aid (FAFSA) for the first time.","FAFSA-to-college handoff unclear, school code not surfaced.","N"),
 ("T09","Pay for it","Complete financial-aid verification before the deadline.","Multi-step, jargon-heavy, deadline not surfaced, easy to abandon.","N"),
 ("T10","Pay for it","Find out when aid disburses and when your refund lands.","Disbursement and refund timing hard to find, no clear calendar.","N"),
 ("T11","Pay for it","Find scholarships you actually qualify for.","Generic scholarship list, eligibility unclear, DACA/F-1 cases unaddressed.","N"),
 ("T12","Pay for it","Set up a payment plan or pay a bill.","Cashier vs aid confusion, payment-plan path buried.","N"),
 ("T13","Pay for it","Get emergency aid fast during a crisis.","Emergency aid exists but unfindable, slow turnaround, no clear ask.","N"),
 # Stage 3 — Plan the path
 ("T14","Plan the path","Find or get assigned an advisor.","No advisor in SIS, students unaware they can get one.","Y"),
 ("T15","Plan the path","Book an advising appointment.","Booking tool hidden or login-walled, long wait, no after-hours option.","Y"),
 ("T16","Plan the path","Choose or change your major / Field of Interest.","FOI not district-wide, undecided students get no scaffolding.","N"),
 ("T17","Plan the path","Get prior or transfer credits evaluated.","Transcript submission unclear, evaluation slow, old credits in limbo.","N"),
 ("T18","Plan the path","Self-enroll in your first classes.","Class search confusing, prereq holds unexplained, no plan to follow.","N"),
 # Stage 4 — Get set up
 ("T19","Get set up","Figure out that Canvas is the LMS and that classes live there.","No one says where classes are, LMS unnamed, login path unclear.","N"),
 ("T20","Get set up","Log in to student email and find official messages.","Two logins, which email is official, where messages go.","N"),
 ("T21","Get set up","Use campus tech: print a form, get on wi-fi, use a lab.","Top frustration: printing, wi-fi, DUO/login, lab access all friction.","N"),
 ("T22","Get set up","Buy or rent course materials / find the bookstore.","Bookstore link buried, which materials for which section unclear.","N"),
 # Stage 5 — In class / fixing enrollment
 ("T23","In class","Add a class after the term has started.","Add deadline and process unclear, instructor-permission step hidden.","N"),
 ("T24","In class","Drop a class and know the refund and withdrawal deadline.","Drop vs withdraw, refund line, deadline all hard to find together.","N"),
 ("T25","In class","Understand how dropping affects your financial aid.","Aid impact of a drop not surfaced at the moment of dropping.","Y"),
 ("T26","In class","Reach your instructor when something goes wrong.","Contact path scattered, no single way to reach faculty.","N"),
 # Stage 6 — When you are struggling
 ("T27","When struggling","Find and use tutoring (in person or online).","Tutoring location/hours/online tool fragmented across pages.","N"),
 ("T28","When struggling","Set up Disability Resources (DRS) accommodations.","DRS intake heavy, documentation steps unclear, no online-only path.","N"),
 ("T29","When struggling","Find counseling / mental-health support and confirm it is confidential.","Counseling vs academic advising conflated, confidentiality unstated.","N"),
 ("T30","When struggling","Find crisis / CARE-BIT support in an urgent moment.","Crisis path not prominent, after-hours coverage unclear.","N"),
 # Stage 7 — When life gets hard (basic needs)
 ("T31","Life gets hard","Find food / a food pantry on campus.","Basic-needs hub exists but unadvertised, eligibility unstated.","N"),
 ("T32","Life gets hard","Find housing, childcare, or basic-needs referrals.","Referrals scattered, no single basic-needs front door.","N"),
 ("T33","Life gets hard","Find the student health clinic or health resources.","Health services hard to locate, hours and cost unclear.","N"),
 # Stage 8 — Belong and grow
 ("T34","Belong & grow","Find a club, community, or identity program that fits you.","Population/identity programs (veterans, AIP, TRIO, LGBTQ+) hard to find.","N"),
 ("T35","Belong & grow","Find a campus job or work-study.","Work-study vs campus job unclear, application path buried.","N"),
 # Stage 9 — Look ahead
 ("T36","Look ahead","Build a transfer plan to a university with no wasted credits.","Transfer pathways and articulation hard to read, no clear plan.","N"),
 ("T37","Look ahead","Find career services and an internship.","Career platform unnamed, internship path unclear.","N"),
 ("T38","Look ahead","Certify VA / veterans education benefits.","VA certification steps scattered, school certifying official hard to reach.","N"),
 # Stage 10 — Finish and leave
 ("T39","Finish & leave","Order transcripts.","Transcript ordering tool buried, cost and timing unclear.","N"),
 ("T40","Finish & leave","Apply for graduation and run a degree/certificate audit.","Audit tool hard to interpret, graduation application steps unclear.","N"),
 ("T41","Finish & leave","Save your email, Drive, and coursework before the account closes.","No one warns that the account closes, offboarding path absent.","N"),
 # Cross-cutting — swirl / consortium
 ("T42","Swirl","Take one needed class at a second college (consortium).","Second application still required, two logins, two portals.","N"),
 ("T43","Swirl","Confirm consortium aid covers a class at the second college.","Consortium aid rules opaque, which office owns it unclear.","Y"),
 ("T44","Swirl","Confirm the swirl credit and residency carry back home.","Residency carryover recently fixed but not communicated, credit-return doubt.","N"),
 # International / visa
 ("T45","International","Keep F-1 visa status valid while registering and paying.","International-student office and visa rules hard to find together.","N"),
 ("T46","International","Get a foreign transcript evaluated for credit.","Foreign credential evaluation path unclear, slow, jargon-heavy.","N"),
]

# A few tasks need the post-Salesforce wave (advising assignment, booking, aid-impact-at-drop, consortium aid)
# already flagged in the task table via the needs_salesforce column.

# =====================================================================
# 4. SCENARIO GENERATION
#    Each persona has a relevant subset of base tasks (their journey),
#    each task is placed at the persona's primary campus. The bank lists
#    the WRITTEN scenarios; the full 500 comes from multiplying base
#    tasks across all 10 colleges (see the Coverage Math tab).
# =====================================================================
# map persona -> primary campus code (first campus token in their list).
# Match leading token so "SMCC" is not swallowed by "MCC", etc.
import re as _re
def primary_campus(camp_field):
    first=_re.split(r"[;,+]",camp_field)[0].strip()
    for code,_ in sorted(COLLEGES,key=lambda c:-len(c[0])):
        if _re.search(r"\b"+_re.escape(code)+r"\b",first,_re.I):
            return code
    if "rio" in first.lower(): return "Rio"
    return "GCC"

# Relevance: which task codes matter most for each persona (their high-stakes journey).
# We give every persona 6-9 tasks so the written bank lands ~120+ scenarios.
REL={
 "P01":["T01","T02","T08","T09","T10","T18","T19","T21"],
 "P02":["T17","T36","T38","T18","T24","T25","T20","T39"],
 "P03":["T16","T28","T29","T34","T18","T19","T26"],
 "P04":["T06","T08","T11","T03","T10","T18"],
 "P05":["T01","T02","T06","T18","T20","T21"],
 "P06":["T17","T03","T18","T39","T36","T21"],
 "P07":["T31","T32","T13","T24","T25","T10"],
 "P08":["T05","T20","T19","T07","T18"],
 "P09":["T34","T27","T15","T18","T21"],
 "P10":["T28","T18","T20","T19","T26","T21"],
 "P11":["T13","T32","T11","T31","T14","T15"],
 "P12":["T45","T12","T11","T18","T20"],
 "P13":["T01","T03","T16","T18","T37"],
 "P14":["T17","T36","T34","T18","T40"],
 "P15":["T02","T26","T18","T12","T21"],
 "P16":["T11","T08","T34","T18"],
 "P17":["T24","T25","T23","T18"],
 "P18":["T42","T43","T44","T17","T39"],
 "P19":["T04","T01","T02","T18","T21"],
 "P20":["T09","T10","T18","T24","T08"],
 "P21":["T28","T38","T29","T18","T15"],
 "P22":["T18","T12","T26","T27","T21","T20"],
 "P23":["T29","T34","T30","T15"],
 "P24":["T10","T08","T22","T18"],
 "P25":["T01","T31","T32","T12","T18"],
 "P26":["T19","T20","T18","T26"],
 "P27":["T42","T43","T44","T20"],
 "P28":["T06","T31","T32","T08","T18"],
 "P29":["T05","T19","T36","T18"],
 "P30":["T06","T18","T27","T07"],
 "P31":["T21","T20","T39","T18"],
 "P32":["T28","T34","T29","T18"],
 "P33":["T38","T04","T18","T21"],
 "P34":["T41","T39","T40","T34"],
 "P35":["T14","T15","T16","T18"],
 "P36":["T46","T45","T17","T18"],
}

PMAP={p[0]:p for p in PERSONAS}
TMAP={t[0]:t for t in TASKS}

scenarios=[]  # rows for the bank
sid=0
for pid,tasklist in REL.items():
    p=PMAP[pid]
    camp=primary_campus(p[8])
    for tc in tasklist:
        t=TMAP[tc]
        sid+=1
        scenario_id=f"S{sid:03d}"
        # alternate human / AI tester across the bank (roughly even split)
        tester_type="Human" if sid%2==1 else "AI"
        scenarios.append({
            "id":scenario_id,
            "persona":f"{pid} {p[1]}",
            "campus":camp,
            "stage":t[1],
            "task":t[2],
            "barrier":t[3],
            "tester_type":tester_type,
            "needs_sf":t[4],
            "status":"Not started",
            "tester":"",
            "date":"",
            "result":"",
            "severity":"",
        })

# =====================================================================
# 5. WRITE THE WORKBOOK
# =====================================================================
wb=openpyxl.Workbook()

def style_header(ws,row,ncols,color=PLUM):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.font=F(10,True,"FFFFFF"); cell.fill=fill(color)
        cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=BORDER

# ---- Tab 0: Read Me ----
ws=wb.active; ws.title="Read Me"
ws.column_dimensions["A"].width=110
readme=[
 ("Maricopa Domain 5 · Student Support & Success — Scenario Bank",18,True,PLUM),
 ("The data backbone for the scaled student-journey usability study.",11,False,INK),
 ("",10,False,INK),
 ("THE MODEL",12,True,SAGE),
 ("Every test is one persona, one campus, one task. The tester (human or AI) opens the tool and receives: this is who you are, these are your barriers, this is the one thing you are trying to do, at this campus. They attempt it from a real felt need.",10,False,INK),
 ("Problems, not destinations. The persona must figure out where to go. We never say go to advising.",10,True,INK),
 ("",10,False,INK),
 ("THE SCALE",12,True,SAGE),
 (f"Persona library: {len(PERSONAS)} personas (extends the original three: Marisol Reyes, Darnell Carter, Alex Nguyen).",10,False,INK),
 (f"Base task taxonomy: {len(TASKS)} tasks across the whole journey, application to offboarding, plus swirl and international tracks.",10,False,INK),
 (f"Written scenarios in this bank: {len(scenarios)} fully specified persona + campus + task + suspected-barrier rows.",10,False,INK),
 ("Coverage math: base tasks multiply by persona and by the 10 colleges. Target is about 500 distinct persona+task scenarios run across all 10 colleges, which is roughly 5,000 task-runs. See the Coverage Math tab.",10,False,INK),
 ("Split: humans run about 250, AI testers run about 250, each set across the 10 colleges. The Tester Type column carries the planned split; flip it as capacity changes.",10,False,INK),
 ("",10,False,INK),
 ("HOW TO USE THIS SHEET",12,True,SAGE),
 ("1. Scenario Bank is the master list. One row per scenario. Filter by Campus, Stage, Tester Type, or Needs Salesforce.",10,False,INK),
 ("2. Assign one scenario to one tester. Put their initials in Tester and the date in Date. Never duplicate a scenario.",10,False,INK),
 ("3. When the run is done, fill Result/Finding and Severity (0 not a usability problem, 1 cosmetic, 2 minor, 3 major, 4 catastrophic) and set Status to Done. Severity combines frequency, impact, and persistence, and a second rater scores it independently.",10,False,INK),
 ("4. The Coverage Dashboard (HTML) reads this structure. Keep the columns as-is.",10,False,INK),
 ("",10,False,INK),
 ("GUARDRAILS",12,True,ROSE),
 ("No PII. Test accounts and initials only. Personas are research instruments, not real people.",10,False,INK),
 ("AI-only scope. Non-AI service or policy fixes get routed to the right owner, not solved here.",10,False,INK),
 ("Translate, don't standardize. Each college keeps its own names. We test findability, not conformity.",10,False,INK),
 ("Human contact is the success metric. Automation takes routine work off staff so they focus on people. No one is replaced.",10,False,INK),
 ("Some scenarios need the Salesforce tool to exist first (advising assignment and booking, aid-impact-at-drop, consortium aid). Those are flagged Needs Salesforce = Y and belong to a later wave.",10,False,INK),
 ("",10,False,INK),
 ("STATUS NOTE",11,True,GOLD),
 ("This sheet is a PLAN and a backbone, not results. The Status, Result, and Severity columns are empty until real runs happen. Nothing here claims a task has been tested.",10,False,INK),
]
r=1
for text,sz,b,color in readme:
    c=ws.cell(row=r,column=1,value=text); c.font=F(sz,b,color); c.alignment=Alignment(wrap_text=True,vertical="top")
    r+=1

# ---- Tab 1: Persona Library ----
ws=wb.create_sheet("Persona Library")
heads=["ID","Name","Age","Background","Barriers","Goal","Tech comfort","Home language","Campus(es)"]
ws.append(heads); style_header(ws,1,len(heads))
for p in PERSONAS:
    ws.append(list(p))
widths=[7,18,6,40,40,38,22,16,22]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(heads)):
    for c in row: c.font=F(9); c.alignment=WRAP; c.border=BORDER
ws.freeze_panes="A2"

# ---- Tab 2: Task Taxonomy ----
ws=wb.create_sheet("Task Taxonomy")
heads=["Task code","Journey stage","Task (problem, not destination)","Suspected barrier","Needs Salesforce?"]
ws.append(heads); style_header(ws,1,len(heads),SAGE)
for t in TASKS:
    ws.append(list(t))
widths=[10,18,46,52,16]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(heads)):
    for c in row: c.font=F(9); c.alignment=WRAP; c.border=BORDER
    if row[4].value=="Y":
        for c in row: c.fill=fill(GOLD_PALE)
ws.freeze_panes="A2"

# ---- Tab 3: Scenario Bank (THE BACKBONE) ----
ws=wb.create_sheet("Scenario Bank")
heads=["Scenario ID","Persona","Campus","Journey Stage","Task","Suspected Barrier",
       "Tester Type","Needs Salesforce?","Status","Tester","Date","Result/Finding","Severity (0 to 4)"]
ws.append(heads); style_header(ws,1,len(heads))
for s in scenarios:
    ws.append([s["id"],s["persona"],s["campus"],s["stage"],s["task"],s["barrier"],
               s["tester_type"],s["needs_sf"],s["status"],s["tester"],s["date"],s["result"],s["severity"]])
widths=[12,20,10,16,42,46,12,14,12,10,12,40,12]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(heads)):
    for c in row: c.font=F(9); c.alignment=WRAP; c.border=BORDER
    if row[7].value=="Y":
        row[7].fill=fill(GOLD_PALE); row[7].font=F(9,True,GOLD)
    row[6].alignment=CTR; row[7].alignment=CTR; row[8].alignment=CTR; row[12].alignment=CTR
ws.freeze_panes="A2"
ws.auto_filter.ref=ws.dimensions

# ---- Tab 4: Coverage Math ----
ws=wb.create_sheet("Coverage Math")
ws.column_dimensions["A"].width=70; ws.column_dimensions["B"].width=18
rows=[
 ("How the bank scales to ~500 scenarios and ~5,000 task-runs","",True,PLUM,13),
 ("","",False,INK,10),
 ("Personas in library",len(PERSONAS),False,INK,10),
 ("Base tasks in taxonomy",len(TASKS),False,INK,10),
 ("Written scenarios in this bank (persona + task)",len(scenarios),False,INK,10),
 ("","",False,INK,10),
 ("Path to ~500 distinct persona+task scenarios","",True,SAGE,11),
 ("Each base task is run by several personas (different barriers reveal different failures).","",False,INK,10),
 ("Roughly 500 = the written persona+task pairs, expanded as the library grows and rarer personas are added per task.","",False,INK,10),
 ("","",False,INK,10),
 ("Path to ~5,000 task-runs","",True,SAGE,11),
 ("~500 distinct scenarios","",False,INK,10),
 ("× 10 colleges (each scenario is attempted at each college, local names differ)","",False,INK,10),
 ("= ~5,000 task-runs","≈ 5,000",True,GOLD,11),
 ("","",False,INK,10),
 ("Human / AI split","",True,SAGE,11),
 ("Humans run ~250 scenarios across the 10 colleges","≈ 2,500 runs",False,INK,10),
 ("AI testers run ~250 scenarios across the 10 colleges","≈ 2,500 runs",False,INK,10),
 ("","",False,INK,10),
 ("Waves","",True,ROSE,11),
 ("Wave 1 (now): everything that can be tested on the live systems today.","",False,INK,10),
 ("Wave 2 (post-Salesforce): scenarios flagged Needs Salesforce = Y (advising assignment + booking, aid-impact-at-drop, consortium aid).","",False,INK,10),
]
r=1
for a,b,bold,color,sz in rows:
    ca=ws.cell(row=r,column=1,value=a); ca.font=F(sz,bold,color); ca.alignment=Alignment(wrap_text=True,vertical="top")
    cb=ws.cell(row=r,column=2,value=b); cb.font=F(sz,bold,color); cb.alignment=CTR
    r+=1

# ---- Tab 5: Accounts We Need ----
ws=wb.create_sheet("Accounts We Need")
heads=["College","Code","Test account status","What we need","Owner to ask"]
ws.append(heads); style_header(ws,1,len(heads),ROSE)
acct_rows=[
 ("Glendale Community College","GCC","Available now (Michelle's own GCC student login)","Run GCC scenarios immediately. Use a no-PII test login where possible.","GCC (in hand)"),
 ("Chandler-Gilbert Community College","CGCC","Needed","One test-student account scoped to CGCC systems.","District / CGCC registrar"),
 ("Estrella Mountain Community College","EMCC","Needed","One test-student account scoped to EMCC systems.","District / EMCC registrar"),
 ("GateWay Community College","GateWay","Needed","One test-student account scoped to GateWay systems.","District / GateWay registrar"),
 ("Mesa Community College","MCC","Needed","One test-student account scoped to MCC systems.","District / MCC registrar"),
 ("Paradise Valley Community College","PVCC","Needed","One test-student account scoped to PVCC systems.","District / PVCC registrar"),
 ("Phoenix College","Phoenix","Needed","One test-student account scoped to Phoenix College systems.","District / Phoenix registrar"),
 ("Rio Salado College","Rio","Needed","One test-student account scoped to Rio Salado systems.","District / Rio registrar"),
 ("Scottsdale Community College","SCC","Needed","One test-student account scoped to SCC systems.","District / SCC registrar"),
 ("South Mountain Community College","SMCC","Needed","One test-student account scoped to SMCC systems.","District / SMCC registrar"),
]
for row in acct_rows: ws.append(list(row))
note=ws.cell(row=ws.max_row+2,column=1,value="District Test Student 1/2/3 accounts (referenced in WORKING-STATE) may be the source. We need confirmation of which test accounts can reach each college's live systems, and whether they can authenticate the advising/booking flows. No PII on any account; test logins and initials only.")
note.font=F(9,False,INK); note.alignment=Alignment(wrap_text=True,vertical="top")
ws.merge_cells(start_row=note.row,start_column=1,end_row=note.row,end_column=5)
widths=[34,10,40,46,24]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=len(acct_rows)+1,max_col=len(heads)):
    for c in row: c.font=F(9); c.alignment=WRAP; c.border=BORDER
    if row[0].value and "Available now" in str(row[2].value): row[2].fill=fill(SAGE_PALE)
    else: row[2].fill=fill(ROSE_PALE)
ws.freeze_panes="A2"

# ---- Tab 6: Barrier Log (compatible with the existing kit) ----
ws=wb.create_sheet("Barrier Log")
heads=["Scenario ID","College","Persona","Task","First search terms used","Path taken",
       "# dead ends","Needed a human? (who)","Where they landed (local name)","Exists here? (Y/N)",
       "Time (min)","Severity (0 to 4)","Barrier description","AI opportunity idea","Tester","Date"]
ws.append(heads); style_header(ws,1,len(heads),GOLD)
widths=[12,12,18,36,28,30,9,18,28,12,9,12,40,40,10,12]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for c in range(1,len(heads)+1):
    ws.cell(row=2,column=c).border=BORDER
ws.freeze_panes="A2"
note=ws.cell(row=ws.max_row+2,column=1,value="One row per completed task-run. Keys back to Scenario ID in the Scenario Bank. Mirrors the original Study Kit Barrier Log so existing tooling still fits. The Jotform capture form (live) feeds rows here.")
note.font=F(9,False,INK); note.alignment=Alignment(wrap_text=True,vertical="top")

wb.save("/sessions/amazing-zealous-turing/mnt/singletrackmom.github.io/airc-sss/Maricopa_Scenario_Bank.xlsx")
print("WROTE Maricopa_Scenario_Bank.xlsx")
print("personas:",len(PERSONAS),"tasks:",len(TASKS),"scenarios:",len(scenarios))
# print campus distribution for the dashboard
from collections import Counter
print("by campus:",dict(Counter(s["campus"] for s in scenarios)))
print("by stage:",dict(Counter(s["stage"] for s in scenarios)))
print("needs SF:",sum(1 for s in scenarios if s["needs_sf"]=="Y"))
print("human:",sum(1 for s in scenarios if s["tester_type"]=="Human"),"AI:",sum(1 for s in scenarios if s["tester_type"]=="AI"))
